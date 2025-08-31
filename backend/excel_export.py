import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from models import Donation, Sponsorship, db
from sqlalchemy import func

class ExcelExporter:
    def __init__(self):
        self.workbook = Workbook()
        self.ws = self.workbook.active
        self.ws.title = "Summary"
        
        # Define styles
        self.header_font = Font(bold=True, color="FFFFFF")
        self.header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        self.subheader_font = Font(bold=True, color="FFFFFF")
        self.subheader_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

    def create_metadata_sheet(self):
        """Create the metadata sheet with summary information"""
        # Get statistics
        stats = self._get_statistics()
        
        # Title
        self.ws['A1'] = "DONATION COLLECTION REPORT"
        self.ws['A1'].font = Font(bold=True, size=16)
        self.ws.merge_cells('A1:H1')
        self.ws['A1'].alignment = Alignment(horizontal='center')
        
        # Metadata section
        row = 3
        self.ws[f'A{row}'] = "METADATA"
        self.ws[f'A{row}'].font = self.header_font
        self.ws[f'A{row}'].fill = self.header_fill
        self.ws.merge_cells(f'A{row}:H{row}')
        
        row += 1
        metadata_data = [
            ["Exported At", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
            ["Last Collection At", stats['last_collection_date']],
            ["Towers Covered", stats['towers_covered']],
            ["Total Donation Amount", f"₹{stats['total_amount']:,.2f}"],
            ["Apartments Visited", stats['apartments_visited']],
            ["Apartments For Follow Up", stats['apartments_follow_up']],
            ["Apartments Remaining", stats['apartments_remaining']],
            ["Total Apartments", stats['total_apartments']]
        ]
        
        for label, value in metadata_data:
            self.ws[f'A{row}'] = label
            self.ws[f'B{row}'] = value
            self.ws[f'A{row}'].font = Font(bold=True)
            self.ws[f'A{row}'].border = self.border
            self.ws[f'B{row}'].border = self.border
            row += 1
        
        # Summary by Tower
        row += 1
        self.ws[f'A{row}'] = "SUMMARY BY TOWER"
        self.ws[f'A{row}'].font = self.header_font
        self.ws[f'A{row}'].fill = self.header_fill
        self.ws.merge_cells(f'A{row}:H{row}')
        
        row += 1
        tower_headers = ["Tower", "Total Amount", "Donations", "Follow-ups", "Skipped", "Remaining", "Completion %"]
        for col, header in enumerate(tower_headers, 1):
            cell = self.ws[f'{get_column_letter(col)}{row}']
            cell.value = header
            cell.font = self.subheader_font
            cell.fill = self.subheader_fill
            cell.border = self.border
            cell.alignment = Alignment(horizontal='center')
        
        row += 1
        for tower in range(1, 11):  # Towers 1-10
            tower_stats = self._get_tower_statistics(tower)
            if tower_stats['total_apartments'] > 0:
                self.ws[f'A{row}'] = f"Tower {chr(64 + tower)}"
                self.ws[f'B{row}'] = f"₹{tower_stats['total_amount']:,.2f}"
                self.ws[f'C{row}'] = tower_stats['donations']
                self.ws[f'D{row}'] = tower_stats['follow_ups']
                self.ws[f'E{row}'] = tower_stats['skipped']
                self.ws[f'F{row}'] = tower_stats['remaining']
                completion_pct = (tower_stats['donations'] + tower_stats['follow_ups'] + tower_stats['skipped']) / tower_stats['total_apartments'] * 100
                self.ws[f'G{row}'] = f"{completion_pct:.1f}%"
                
                for col in range(1, 8):
                    cell = self.ws[f'{get_column_letter(col)}{row}']
                    cell.border = self.border
                    if col == 2:  # Amount column
                        cell.number_format = '#,##0.00'
                    elif col == 7:  # Percentage column
                        cell.number_format = '0.0%'
                row += 1
        
        # Auto-adjust column widths
        for column in self.ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            self.ws.column_dimensions[column_letter].width = adjusted_width

    def create_sponsorship_sheet(self):
        """Create the sponsorship summary sheet with filtering"""
        # Create new sheet
        ws = self.workbook.create_sheet("Sponsorship Summary")
        
        # Title
        ws['A1'] = "SPONSORSHIP SUMMARY REPORT"
        ws['A1'].font = Font(bold=True, size=16)
        ws.merge_cells('A1:H1')
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Get sponsorship statistics
        sponsorship_stats = self._get_sponsorship_statistics()
        
        # Metadata section
        row = 3
        ws[f'A{row}'] = "SPONSORSHIP METADATA"
        ws[f'A{row}'].font = self.header_font
        ws[f'A{row}'].fill = self.header_fill
        ws.merge_cells(f'A{row}:H{row}')
        
        row += 1
        metadata_data = [
            ["Exported At", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
            ["Total Sponsorships", sponsorship_stats['total_sponsorships']],
            # ["Total Sponsorship Value", f"₹{sponsorship_stats['total_value']:,.2f}"],
            ["Booked Sponsorships", sponsorship_stats['booked_sponsorships']],
            ["Available Sponsorships", sponsorship_stats['available_sponsorships']],
            ["Total Bookings", sponsorship_stats['total_bookings']],
            # ["Total Max Capacity", sponsorship_stats['total_max_capacity']],
            ["Utilization Rate", f"{sponsorship_stats['utilization_rate']:.1f}%"]
        ]
        
        for label, value in metadata_data:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'A{row}'].border = self.border
            ws[f'B{row}'].border = self.border
            row += 1
        
        # Sponsorship Details Table
        row += 2
        ws[f'A{row}'] = "SPONSORSHIP DETAILS"
        ws[f'A{row}'].font = self.header_font
        ws[f'A{row}'].fill = self.header_fill
        ws.merge_cells(f'A{row}:H{row}')
        
        row += 1
        headers = [
            "ID", "Sponsorship Name", "Amount", "Max Count", 
            "Booked Count", "Available", "Status", "Created Date"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws[f'{get_column_letter(col)}{row}']
            cell.value = header
            cell.font = self.subheader_font
            cell.fill = self.subheader_fill
            cell.border = self.border
            cell.alignment = Alignment(horizontal='center')
        
        # Enable filters for the header row
        ws.auto_filter.ref = f"A{row}:{get_column_letter(len(headers))}{row}"
        
        # Get all sponsorships
        sponsorships = Sponsorship.query.all()
        
        row += 1
        for sponsorship in sponsorships:
            available = sponsorship.max_count - sponsorship.booked
            status = "Closed" if sponsorship.is_closed else "Available" if available > 0 else "Partially Booked"
            
            ws[f'A{row}'] = sponsorship.id
            ws[f'B{row}'] = sponsorship.name
            ws[f'C{row}'] = float(sponsorship.amount)
            ws[f'D{row}'] = sponsorship.max_count
            ws[f'E{row}'] = sponsorship.booked
            ws[f'F{row}'] = available
            ws[f'G{row}'] = status
            ws[f'H{row}'] = sponsorship.created_at.strftime("%Y-%m-%d %H:%M")
            
            # Apply borders and formatting
            for col in range(1, 9):
                cell = ws[f'{get_column_letter(col)}{row}']
                cell.border = self.border
                if col == 3:  # Amount column
                    cell.number_format = '₹#,##0.00'
                elif col == 4 or col == 5 or col == 6:  # Count columns
                    cell.number_format = '0'
            
            row += 1
        
        # Sponsorship Donations Table
        row += 2
        ws[f'A{row}'] = "DONATIONS WITH SPONSORSHIP"
        ws[f'A{row}'].font = self.header_font
        ws[f'A{row}'].fill = self.header_fill
        ws.merge_cells(f'A{row}:K{row}')
        
        row += 1
        donation_headers = [
            "Donation ID", "Tower", "Apartment", "Donor Name", "Amount", 
            "Sponsorship Name", "Phone Number", "Head Count", "UPI/Other Person", 
            "Notes", "Date"
        ]
        
        for col, header in enumerate(donation_headers, 1):
            cell = ws[f'{get_column_letter(col)}{row}']
            cell.value = header
            cell.font = self.subheader_font
            cell.fill = self.subheader_fill
            cell.border = self.border
            cell.alignment = Alignment(horizontal='center')
        
        # Enable filters for the donation header row
        ws.auto_filter.ref = f"A{row}:{get_column_letter(len(donation_headers))}{row}"
        
        # Get donations with sponsorship
        donations_with_sponsorship = Donation.query.filter(
            Donation.sponsorship.isnot(None),
            Donation.sponsorship != ""
        ).order_by(Donation.created_at.desc()).all()
        
        row += 1
        for donation in donations_with_sponsorship:
            apartment = f"{chr(64 + donation.tower)}{donation.floor}{donation.unit:02d}"
            
            ws[f'A{row}'] = donation.id
            ws[f'B{row}'] = f"Tower {chr(64 + donation.tower)}"
            ws[f'C{row}'] = apartment
            ws[f'D{row}'] = donation.donor_name
            ws[f'E{row}'] = float(donation.amount)
            ws[f'F{row}'] = donation.sponsorship or ""
            ws[f'G{row}'] = donation.phone_number or ""
            ws[f'H{row}'] = donation.head_count or ""
            ws[f'I{row}'] = donation.upi_other_person or ""
            ws[f'J{row}'] = donation.notes or ""
            ws[f'K{row}'] = donation.created_at.strftime("%Y-%m-%d %H:%M")
            
            # Apply borders and formatting
            for col in range(1, 12):
                cell = ws[f'{get_column_letter(col)}{row}']
                cell.border = self.border
                if col == 5:  # Amount column
                    cell.number_format = '₹#,##0.00'
                elif col == 8:  # Head count column
                    cell.number_format = '0'
            
            row += 1
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            ws.column_dimensions[column_letter].width = adjusted_width

    def create_tower_sheets(self):
        """Create individual sheets for each tower"""
        sheets_created = 0
        for tower in range(1, 11):  # Towers 1-10
            tower_donations = self._get_tower_donations(tower)
            if not tower_donations:
                continue
                
            # Create new sheet
            ws = self.workbook.create_sheet(f"Tower {chr(64 + tower)}")
            sheets_created += 1
            
            # Title
            ws['A1'] = f"TOWER {chr(64 + tower)} - DONATION DETAILS"
            ws['A1'].font = Font(bold=True, size=14)
            ws.merge_cells('A1:K1')
            ws['A1'].alignment = Alignment(horizontal='center')
            
            # Headers
            row = 3
            headers = [
                "Floor", "Unit", "Apartment", "Donor Name", "Amount", 
                "Phone Number", "Head Count", "UPI/Other Person", 
                "Sponsorship", "Notes", "Status", "Date"
            ]
            
            for col, header in enumerate(headers, 1):
                cell = ws[f'{get_column_letter(col)}{row}']
                cell.value = header
                cell.font = self.header_font
                cell.fill = self.header_fill
                cell.border = self.border
                cell.alignment = Alignment(horizontal='center')
            
            # Enable filters for the header row
            ws.auto_filter.ref = f"A{row}:{get_column_letter(len(headers))}{row}"
            
            # Data
            row += 1
            for donation in tower_donations:
                apartment = f"{chr(64 + tower)}{donation.floor}{donation.unit:02d}"
                
                ws[f'A{row}'] = donation.floor
                ws[f'B{row}'] = donation.unit
                ws[f'C{row}'] = apartment
                ws[f'D{row}'] = donation.donor_name
                ws[f'E{row}'] = donation.amount
                ws[f'F{row}'] = donation.phone_number or ""
                ws[f'G{row}'] = donation.head_count or ""
                ws[f'H{row}'] = donation.upi_other_person or ""
                ws[f'I{row}'] = donation.sponsorship or ""
                ws[f'J{row}'] = donation.notes or ""
                ws[f'K{row}'] = donation.status
                ws[f'L{row}'] = donation.created_at.strftime("%Y-%m-%d %H:%M")
                
                # Apply borders and formatting
                for col in range(1, 13):
                    cell = ws[f'{get_column_letter(col)}{row}']
                    cell.border = self.border
                    if col == 5:  # Amount column
                        cell.number_format = '#,##0.00'
                
                row += 1
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 30)
                ws.column_dimensions[column_letter].width = adjusted_width

    def _get_sponsorship_statistics(self):
        """Get sponsorship statistics for the metadata section"""
        # Get all sponsorships
        sponsorships = Sponsorship.query.all()
        
        total_sponsorships = len(sponsorships)
        total_value = sum(float(s.amount) * s.max_count for s in sponsorships)
        booked_sponsorships = sum(1 for s in sponsorships if s.is_closed)
        available_sponsorships = sum(1 for s in sponsorships if not s.is_closed)
        total_bookings = sum(s.booked for s in sponsorships)
        total_max_capacity = sum(s.max_count for s in sponsorships)
        
        utilization_rate = (total_bookings / total_max_capacity * 100) if total_max_capacity > 0 else 0
        
        return {
            'total_sponsorships': total_sponsorships,
            'total_value': total_value,
            'booked_sponsorships': booked_sponsorships,
            'available_sponsorships': available_sponsorships,
            'total_bookings': total_bookings,
            'total_max_capacity': total_max_capacity,
            'utilization_rate': utilization_rate
        }

    def _get_statistics(self):
        """Get overall statistics for the metadata sheet"""
        # Total amount
        total_amount = db.session.query(func.sum(Donation.amount)).filter_by(status='completed').scalar() or 0
        
        # Last collection date
        last_donation = Donation.query.order_by(Donation.created_at.desc()).first()
        last_collection_date = last_donation.created_at.strftime("%Y-%m-%d %H:%M:%S") if last_donation else "No donations yet"
        
        # Counts by status
        completed_count = Donation.query.filter_by(status='completed').count()
        follow_up_count = Donation.query.filter_by(status='follow-up').count()
        skipped_count = Donation.query.filter_by(status='skipped').count()
        
        # Total apartments (14 floors * 4 units * 10 towers)
        total_apartments = 14 * 4 * 10
        apartments_visited = completed_count + follow_up_count + skipped_count
        apartments_remaining = total_apartments - apartments_visited
        
        # Towers covered (towers with at least one donation)
        towers_covered = db.session.query(func.count(func.distinct(Donation.tower))).scalar() or 0
        
        return {
            'total_amount': float(total_amount),
            'last_collection_date': last_collection_date,
            'towers_covered': towers_covered,
            'apartments_visited': apartments_visited,
            'apartments_follow_up': follow_up_count,
            'apartments_remaining': apartments_remaining,
            'total_apartments': total_apartments
        }

    def _get_tower_statistics(self, tower):
        """Get statistics for a specific tower"""
        # Total apartments in tower
        total_apartments = 14 * 4  # 14 floors * 4 units
        
        # Counts by status for this tower
        donations = Donation.query.filter_by(tower=tower, status='completed').count()
        follow_ups = Donation.query.filter_by(tower=tower, status='follow-up').count()
        skipped = Donation.query.filter_by(tower=tower, status='skipped').count()
        
        # Total amount for this tower
        total_amount = db.session.query(func.sum(Donation.amount)).filter_by(tower=tower, status='completed').scalar() or 0
        
        # Remaining apartments
        remaining = total_apartments - (donations + follow_ups + skipped)
        
        return {
            'total_apartments': total_apartments,
            'donations': donations,
            'follow_ups': follow_ups,
            'skipped': skipped,
            'remaining': remaining,
            'total_amount': float(total_amount)
        }

    def _get_tower_donations(self, tower):
        """Get all donations for a specific tower"""
        return Donation.query.filter_by(tower=tower).order_by(Donation.floor.desc(), Donation.unit).all()

    def generate_excel(self):
        """Generate the complete Excel file"""
        try:
            self.create_metadata_sheet()
            self.create_sponsorship_sheet()
            self.create_tower_sheets()
            
            # Remove the default sheet if it exists and we have other sheets
            if 'Sheet' in self.workbook.sheetnames and len(self.workbook.sheetnames) > 1:
                self.workbook.remove(self.workbook['Sheet'])
            
            # Ensure we have at least one sheet
            if len(self.workbook.sheetnames) == 0:
                raise Exception("No sheets were created for the Excel file")
            
            # Save to bytes
            excel_file = io.BytesIO()
            self.workbook.save(excel_file)
            excel_file.seek(0)
            
            return excel_file
        except Exception as e:
            # Log the error for debugging
            print(f"Error generating Excel file: {str(e)}")
            raise

def export_donations_to_excel():
    """Main function to export donations to Excel"""
    exporter = ExcelExporter()
    return exporter.generate_excel()
